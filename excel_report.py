import json
from pathlib import Path
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


COLOR_MAP = {
    "red": "FF4D4D",
    "yellow": "FFD966",
    "green": "A9D18E",
    "blue": "9DC3E6",
    "orange": "F4B183",
    "purple": "B4A7D6",
    "grey": "D9EAD3",
    "gray": "D9EAD3",
}


def to_decimal(value):
    try:
        return Decimal(str(value or "0").replace(",", ""))
    except Exception:
        return Decimal("0.00")


def money(value):
    return float(to_decimal(value))


def get_fill(color_name):
    color = COLOR_MAP.get(str(color_name).lower(), "D9EAD3")
    return PatternFill("solid", fgColor=color)


def header_fill():
    return PatternFill("solid", fgColor="5B9BD5")


def dark_fill():
    return PatternFill("solid", fgColor="1F4E79")


def light_blue_fill():
    return PatternFill("solid", fgColor="DDEBF7")


def white_font():
    return Font(color="FFFFFF", bold=True)


def bold_font():
    return Font(bold=True)


def title_font():
    return Font(size=16, bold=True)


def thin_border():
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_table_style(sheet, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize_columns(sheet, min_width=10, max_width=42):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        sheet.column_dimensions[column_letter].width = adjusted_width


def create_excel_report(
    processed_transactions,
    summary,
    config,
    output_path,
    learned_rules_path=None
):
    wb = Workbook()

    dashboard = wb.active
    dashboard.title = "Budget Dashboard"

    create_dashboard_sheet(dashboard, processed_transactions, summary, config)

    category_sheet = wb.create_sheet("Category Summary")
    create_category_summary_sheet(category_sheet, summary, config)

    tx_sheet = wb.create_sheet("All Transactions")
    create_transactions_sheet(tx_sheet, processed_transactions)

    review_sheet = wb.create_sheet("Needs Review")
    create_needs_review_sheet(review_sheet, processed_transactions)

    transfer_sheet = wb.create_sheet("Transfers")
    create_transfers_sheet(transfer_sheet, processed_transactions)

    rules_sheet = wb.create_sheet("Learned Rules")
    create_learned_rules_sheet(rules_sheet, learned_rules_path)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)


def create_dashboard_sheet(sheet, processed_transactions, summary, config):
    sheet["A1"] = "Budget Report"
    sheet["A1"].font = title_font()

    sheet["A2"] = f"{summary.get('budget_start_date')} to {summary.get('budget_end_date')}"
    sheet["A2"].font = Font(italic=True)

    sheet["A4"] = "Main Summary"
    sheet["A4"].fill = dark_fill()
    sheet["A4"].font = white_font()
    sheet.merge_cells("A4:D4")

    summary_rows = [
        ("Total Income", money(summary.get("total_income"))),
        ("Category Expenses", money(summary.get("total_category_expenses", summary.get("total_expenses")))),
        ("Transfers to Other Accounts", money(summary.get("total_external_transfers_to_other_accounts"))),
        ("Transfers Included as Expense", money(summary.get("total_external_transfers_to_other_accounts_included_as_expense"))),
        ("Transfers Excluded from Expense", money(summary.get("total_external_transfers_to_other_accounts_excluded_from_expense"))),
        ("Total Expenses Used for Savings", money(summary.get("total_expenses"))),
        ("Planned Savings", money(summary.get("planned_savings"))),
        ("Actual Savings", money(summary.get("actual_savings"))),
        ("Savings Difference", money(summary.get("savings_difference"))),
        ("Internal Transfers Ignored", money(summary.get("total_internal_transfers_ignored_from_spending"))),
        ("Transactions Needing Review", summary.get("review_count", 0)),
    ]

    start_row = 5

    for index, (label, value) in enumerate(summary_rows, start=start_row):
        sheet.cell(row=index, column=1).value = label
        sheet.cell(row=index, column=2).value = value

        sheet.cell(row=index, column=1).font = bold_font()

        if "Review" in label:
            sheet.cell(row=index, column=2).number_format = '0'
        else:
            sheet.cell(row=index, column=2).number_format = '€#,##0.00'

    savings_difference = money(summary.get("savings_difference"))

    status_row = start_row + len(summary_rows) + 1

    sheet.cell(row=status_row, column=1).value = "Saved well or over budget?"
    sheet.cell(row=status_row, column=1).font = bold_font()

    if savings_difference >= 0:
        sheet.cell(row=status_row, column=2).value = "SAVED WELL"
        sheet.cell(row=status_row, column=2).fill = PatternFill("solid", fgColor="A9D18E")
    else:
        sheet.cell(row=status_row, column=2).value = "OVER BUDGET"
        sheet.cell(row=status_row, column=2).fill = PatternFill("solid", fgColor="FF4D4D")

    sheet.cell(row=status_row, column=2).font = bold_font()

    current_row = status_row + 3

    sheet.cell(row=current_row, column=1).value = "Planned vs Reality by Category"
    sheet.cell(row=current_row, column=1).fill = header_fill()
    sheet.cell(row=current_row, column=1).font = white_font()
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

    current_row += 1

    headers = ["Category", "Color", "Planned Limit", "Actual Spend", "Remaining"]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        cell.fill = light_blue_fill()
        cell.font = bold_font()

    current_row += 1

    comparison = summary.get("category_limit_comparison", [])

    for item in comparison:
        category_name = item.get("category_name", "")
        category = find_category(config, category_name)
        color = category.get("color", "") if category else ""

        sheet.cell(row=current_row, column=1).value = category_name
        sheet.cell(row=current_row, column=2).value = color
        sheet.cell(row=current_row, column=3).value = money(item.get("planned_limit"))
        sheet.cell(row=current_row, column=4).value = money(item.get("actual_spend"))
        sheet.cell(row=current_row, column=5).value = money(item.get("remaining"))

        if color:
            sheet.cell(row=current_row, column=1).fill = get_fill(color)

        for col in [3, 4, 5]:
            sheet.cell(row=current_row, column=col).number_format = '€#,##0.00'

        if money(item.get("remaining")) < 0:
            sheet.cell(row=current_row, column=5).fill = PatternFill("solid", fgColor="FF4D4D")
        else:
            sheet.cell(row=current_row, column=5).fill = PatternFill("solid", fgColor="A9D18E")

        current_row += 1

    current_row += 2

    sheet.cell(row=current_row, column=1).value = "Planned vs Reality by Subcategory"
    sheet.cell(row=current_row, column=1).fill = header_fill()
    sheet.cell(row=current_row, column=1).font = white_font()
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)

    current_row += 1

    headers = ["Category", "Subcategory", "Planned Limit", "Actual Spend", "Remaining", "Status"]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        cell.fill = light_blue_fill()
        cell.font = bold_font()

    current_row += 1

    for item in summary.get("subcategory_limit_comparison", []):
        sheet.cell(row=current_row, column=1).value = item.get("category_name", "")
        sheet.cell(row=current_row, column=2).value = item.get("subcategory_name", "")

        if item.get("planned_limit") != "":
            sheet.cell(row=current_row, column=3).value = money(item.get("planned_limit"))
            sheet.cell(row=current_row, column=3).number_format = '€#,##0.00'
        else:
            sheet.cell(row=current_row, column=3).value = ""

        sheet.cell(row=current_row, column=4).value = money(item.get("actual_spend"))
        sheet.cell(row=current_row, column=4).number_format = '€#,##0.00'

        if item.get("remaining") != "":
            sheet.cell(row=current_row, column=5).value = money(item.get("remaining"))
            sheet.cell(row=current_row, column=5).number_format = '€#,##0.00'
        else:
            sheet.cell(row=current_row, column=5).value = ""

        sheet.cell(row=current_row, column=6).value = item.get("status", "")

        if item.get("status") == "over_limit":
            sheet.cell(row=current_row, column=5).fill = PatternFill("solid", fgColor="FF4D4D")
        elif item.get("status") == "within_limit":
            sheet.cell(row=current_row, column=5).fill = PatternFill("solid", fgColor="A9D18E")

        current_row += 1

    apply_table_style(sheet, 4, current_row, 1, 6)
    autosize_columns(sheet)

    sheet.freeze_panes = "A5"


def create_category_summary_sheet(sheet, summary, config):
    sheet["A1"] = "Category Summary"
    sheet["A1"].font = title_font()

    headers = ["Category", "Subcategory", "Planned Limit", "Actual Spend", "Remaining", "Status"]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill()
        cell.font = white_font()

    row = 4

    for item in summary.get("subcategory_limit_comparison", []):
        sheet.cell(row=row, column=1).value = item.get("category_name", "")
        sheet.cell(row=row, column=2).value = item.get("subcategory_name", "")

        if item.get("planned_limit") != "":
            sheet.cell(row=row, column=3).value = money(item.get("planned_limit"))
            sheet.cell(row=row, column=3).number_format = '€#,##0.00'

        sheet.cell(row=row, column=4).value = money(item.get("actual_spend"))
        sheet.cell(row=row, column=4).number_format = '€#,##0.00'

        if item.get("remaining") != "":
            sheet.cell(row=row, column=5).value = money(item.get("remaining"))
            sheet.cell(row=row, column=5).number_format = '€#,##0.00'

        sheet.cell(row=row, column=6).value = item.get("status", "")

        row += 1

    row += 1

    sheet.cell(row=row, column=1).value = "Transfers to Other Accounts"
    sheet.cell(row=row, column=4).value = money(summary.get("total_external_transfers_to_other_accounts"))
    sheet.cell(row=row, column=4).number_format = '€#,##0.00'
    sheet.cell(row=row, column=1).font = bold_font()

    row += 1

    sheet.cell(row=row, column=1).value = "Transfers Included as Expense"
    sheet.cell(row=row, column=4).value = money(summary.get("total_external_transfers_to_other_accounts_included_as_expense"))
    sheet.cell(row=row, column=4).number_format = '€#,##0.00'

    row += 1

    sheet.cell(row=row, column=1).value = "Transfers Excluded from Expense"
    sheet.cell(row=row, column=4).value = money(summary.get("total_external_transfers_to_other_accounts_excluded_from_expense"))
    sheet.cell(row=row, column=4).number_format = '€#,##0.00'

    apply_table_style(sheet, 3, row, 1, 6)
    autosize_columns(sheet)
    sheet.freeze_panes = "A4"


def create_transactions_sheet(sheet, transactions):
    sheet["A1"] = "All Transactions"
    sheet["A1"].font = title_font()

    headers = [
        "date",
        "account_name",
        "description",
        "debit",
        "credit",
        "amount",
        "transaction_type",
        "category_name",
        "subcategory",
        "include_in_expenses",
        "review_decision",
        "categorization_method",
        "ai_confidence",
        "transfer_match_status",
        "needs_review",
        "review_reasons"
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill()
        cell.font = white_font()

    row = 4

    for tx in transactions:
        for col, header in enumerate(headers, start=1):
            value = tx.get(header, "")

            if header in ["debit", "credit", "amount"]:
                value = money(value)

            sheet.cell(row=row, column=col).value = value

            if header in ["debit", "credit", "amount"]:
                sheet.cell(row=row, column=col).number_format = '€#,##0.00'

        if tx.get("needs_review") == "yes":
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FCE4D6")

        row += 1

    apply_table_style(sheet, 3, row, 1, len(headers))
    autosize_columns(sheet)
    sheet.freeze_panes = "A4"


def create_needs_review_sheet(sheet, transactions):
    sheet["A1"] = "Needs Review"
    sheet["A1"].font = title_font()

    review_transactions = [
        tx for tx in transactions
        if tx.get("needs_review") == "yes"
    ]

    headers = [
        "date",
        "account_name",
        "description",
        "amount",
        "transaction_type",
        "category_name",
        "subcategory",
        "include_in_expenses",
        "review_decision",
        "categorization_method",
        "ai_confidence",
        "review_reasons"
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor="F4B183")
        cell.font = bold_font()

    row = 4

    for tx in review_transactions:
        for col, header in enumerate(headers, start=1):
            value = tx.get(header, "")

            if header == "amount":
                value = money(value)

            sheet.cell(row=row, column=col).value = value

            if header == "amount":
                sheet.cell(row=row, column=col).number_format = '€#,##0.00'

        row += 1

    apply_table_style(sheet, 3, row, 1, len(headers))
    autosize_columns(sheet)
    sheet.freeze_panes = "A4"


def create_transfers_sheet(sheet, transactions):
    sheet["A1"] = "Transfers"
    sheet["A1"].font = title_font()

    transfer_transactions = [
        tx for tx in transactions
        if "transfer" in str(tx.get("transaction_type", "")).lower()
    ]

    headers = [
        "date",
        "account_name",
        "description",
        "amount",
        "transaction_type",
        "category_name",
        "subcategory",
        "include_in_expenses",
        "review_decision",
        "transfer_group_id",
        "transfer_match_status",
        "transfer_counterparty_account"
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.fill = light_blue_fill()
        cell.font = bold_font()

    row = 4

    for tx in transfer_transactions:
        for col, header in enumerate(headers, start=1):
            value = tx.get(header, "")

            if header == "amount":
                value = money(value)

            sheet.cell(row=row, column=col).value = value

            if header == "amount":
                sheet.cell(row=row, column=col).number_format = '€#,##0.00'

        row += 1

    apply_table_style(sheet, 3, row, 1, len(headers))
    autosize_columns(sheet)
    sheet.freeze_panes = "A4"


def create_learned_rules_sheet(sheet, learned_rules_path):
    sheet["A1"] = "Learned Rules"
    sheet["A1"].font = title_font()

    rules = []

    if learned_rules_path and Path(learned_rules_path).exists():
        try:
            with open(learned_rules_path, "r", encoding="utf-8") as file:
                rules = json.load(file)
        except Exception:
            rules = []

    headers = [
        "merchant_pattern",
        "category_id",
        "category_name",
        "subcategory"
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill()
        cell.font = white_font()

    row = 4

    for rule in rules:
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=col).value = rule.get(header, "")

        row += 1

    apply_table_style(sheet, 3, max(row, 4), 1, len(headers))
    autosize_columns(sheet)


def find_category(config, category_name):
    for category in config.get("categories", []):
        if str(category.get("name", "")).lower() == str(category_name).lower():
            return category

    return None